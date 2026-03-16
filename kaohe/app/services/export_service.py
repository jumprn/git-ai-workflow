import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func

from app.extensions import db
from app.models import CommitAIStatus, Project, Member


HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def export_by_project(project_id: int = None, member_id: int = None,
                      start_date: date = None, end_date: date = None) -> io.BytesIO:
    """Export format 1: member | project | AI lines | total lines | coverage rate."""
    records = _query_records(project_id, member_id, start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = 'AI代码覆盖率(按项目)'

    headers = ['成员', '项目', 'AI代码行数', '总代码行数', '覆盖率(%)']
    _write_header(ws, headers)

    for i, r in enumerate(records, start=2):
        ws.cell(row=i, column=1, value=r['member_name'])
        ws.cell(row=i, column=2, value=r['project_name'])
        ws.cell(row=i, column=3, value=r['ai_lines'])
        ws.cell(row=i, column=4, value=r['total_lines'])
        ws.cell(row=i, column=5, value=r['coverage_rate'])
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_by_member(project_id: int = None, member_id: int = None,
                     start_date: date = None, end_date: date = None) -> io.BytesIO:
    """Export format 2: member | AI lines | total lines | coverage rate (aggregated)."""
    query = (
        db.session.query(
            Member.name.label('member_name'),
            func.sum(CommitAIStatus.ai_lines_added).label('ai_lines'),
            func.sum(CommitAIStatus.total_lines).label('total_lines'),
        )
        .join(Project, CommitAIStatus.project_id == Project.id)
        .join(Member, CommitAIStatus.member_id == Member.id)
        .filter(
            Project.is_deleted.is_(False),
            Member.is_deleted.is_(False),
            CommitAIStatus.member_id.isnot(None),
        )
    )

    if project_id:
        query = query.filter(CommitAIStatus.project_id == project_id)
    if member_id:
        query = query.filter(CommitAIStatus.member_id == member_id)
    if start_date:
        query = query.filter(CommitAIStatus.commit_date >= start_date)
    if end_date:
        query = query.filter(CommitAIStatus.commit_date <= end_date)

    results = query.group_by(Member.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'AI代码覆盖率(按成员)'

    headers = ['成员', 'AI代码行数', '总代码行数', '覆盖率(%)']
    _write_header(ws, headers)

    for i, r in enumerate(results, start=2):
        ai = r.ai_lines or 0
        total = r.total_lines or 0
        rate = round(ai / total * 100, 2) if total > 0 else 0
        ws.cell(row=i, column=1, value=r.member_name)
        ws.cell(row=i, column=2, value=ai)
        ws.cell(row=i, column=3, value=total)
        ws.cell(row=i, column=4, value=rate)
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _query_records(project_id, member_id, start_date, end_date) -> list[dict]:
    """Query coverage grouped by member + project from commit_ai_status."""
    query = (
        db.session.query(
            Member.name.label('member_name'),
            Project.name.label('project_name'),
            func.sum(CommitAIStatus.ai_lines_added).label('ai_lines'),
            func.sum(CommitAIStatus.total_lines).label('total_lines'),
        )
        .join(Project, CommitAIStatus.project_id == Project.id)
        .join(Member, CommitAIStatus.member_id == Member.id)
        .filter(
            Project.is_deleted.is_(False),
            Member.is_deleted.is_(False),
            CommitAIStatus.member_id.isnot(None),
        )
    )

    if project_id:
        query = query.filter(CommitAIStatus.project_id == project_id)
    if member_id:
        query = query.filter(CommitAIStatus.member_id == member_id)
    if start_date:
        query = query.filter(CommitAIStatus.commit_date >= start_date)
    if end_date:
        query = query.filter(CommitAIStatus.commit_date <= end_date)

    results = query.group_by(Member.name, Project.name).all()
    return [
        {
            'member_name': r.member_name,
            'project_name': r.project_name,
            'ai_lines': r.ai_lines or 0,
            'total_lines': r.total_lines or 0,
            'coverage_rate': round((r.ai_lines or 0) / r.total_lines * 100, 2) if r.total_lines else 0,
        }
        for r in results
    ]


def _write_header(ws, headers: list[str]):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
